# Minimal Flask app for CORS testing
import os
import time
import datetime
import io
import zipfile
import pandas as pd
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from gmap_robot import GoogleMapsRobot
try:
    from supabase_client import supabase
    from services.material_service import MaterialService
    print("✅ Database connection available")
    material_service = MaterialService(supabase) if supabase else None
except ImportError as e:
    print(f"⚠️ Database modules not available: {e}")
    supabase = None
    material_service = None

app = Flask(__name__)
CORS(app, 
     origins=['http://localhost:5173', 'http://localhost:5174', 'http://127.0.0.1:5173', 'http://127.0.0.1:5174'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization'])

# Mock the Google Maps route for testing
@app.route('/api/gmap/process', methods=['POST'])
def gmap_process():
    data = request.get_json()
    origin = data.get('origin')
    destinations_text = data.get('destinations', '')
    
    if not origin or not destinations_text:
        return jsonify({"error": "必須提供出發地和目的地。"}), 400
    
    destinations = [addr.strip() for addr in destinations_text.split('\n') if addr.strip()]
    session_id = f"session_{int(time.time())}"
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
    
    # Create screenshot folder
    image_folder_path = os.path.join(SCREENSHOTS_FOLDER, today_str, session_id)
    os.makedirs(image_folder_path, exist_ok=True)
    
    try:
        # Use Google Maps robot to get real screenshots
        print(f"Starting Google Maps robot for session {session_id}")
        robot = GoogleMapsRobot(headless=True)
        robot_results = robot.process_multiple_routes(origin, destinations, image_folder_path)
        
        # Convert results format to match frontend expectations
        results = []
        for robot_result in robot_results:
            screenshot_url = ""
            if "image_filename" in robot_result:
                screenshot_url = f"screenshots/{today_str}/{session_id}/{robot_result['image_filename']}"
            
            results.append({
                "origin": robot_result["origin"],
                "destination": robot_result["destination"],
                "distance": robot_result["distance"],
                "image_filename": robot_result.get("image_filename", ""),
                "image_local_path": robot_result.get("image_local_path", ""),
                "screenshot_url": screenshot_url
            })
        
        # Store results in cache for downloads
        SESSION_RESULTS_CACHE[session_id] = results
        
        print(f"✅ Google Maps processing completed for {len(results)} destinations")
        return jsonify({
            "results": results,
            "session_id": session_id
        })
        
    except Exception as e:
        print(f"❌ Google Maps robot error: {e}")
        return jsonify({"error": f"處理時發生錯誤: {str(e)}"}), 500

# Session storage and configuration
SESSION_RESULTS_CACHE = {}

# Path configuration  
basedir = os.path.abspath(os.path.dirname(__file__))
SCREENSHOTS_FOLDER = os.path.join(basedir, 'screenshots')

@app.route('/')
def index():
    return jsonify({
        "message": "MFish Station Backend API (Test Mode)",
        "status": "running",
        "note": "This is a minimal test version to verify CORS configuration"
    })

@app.route('/screenshots/<path:path>')
def send_screenshot(path):
    """Serve real screenshot images from the screenshots folder"""
    try:
        return send_from_directory(SCREENSHOTS_FOLDER, path)
    except Exception as e:
        print(f"Error serving screenshot: {e}")
        return "Screenshot not found", 404

@app.route('/api/download/excel/<session_id>', methods=['GET'])
def download_excel(session_id):
    """Download Excel report for session"""
    session_data = SESSION_RESULTS_CACHE.get(session_id)
    if not session_data:
        return "Session not found or expired.", 404
    
    # Create mock Excel data
    export_data = [
        {
            "起始點": item['origin'], 
            "終點": item['destination'], 
            "距離": item['distance'], 
            "圖片名稱": item['image_filename']
        } for item in session_data
    ]
    
    df = pd.DataFrame(export_data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='路線距離報告')
    
    output.seek(0)
    return send_file(
        output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, 
        download_name=f'google_maps_report_{session_id}.xlsx'
    )

@app.route('/api/download/zip/<session_id>', methods=['GET'])
def download_zip(session_id):
    """Download ZIP file with images for session"""
    session_data = SESSION_RESULTS_CACHE.get(session_id)
    if not session_data:
        return "Session not found or expired.", 404
    
    # Create mock ZIP file
    memory_file = io.BytesIO()
    
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in session_data:
            # Add mock file content to ZIP
            mock_content = f"Mock image content for {item['destination']}"
            zf.writestr(item['image_filename'], mock_content)
    
    memory_file.seek(0)
    return send_file(
        memory_file, 
        mimetype='application/zip', 
        as_attachment=True, 
        download_name=f'map_images_{session_id}.zip'
    )

# Materials API endpoints
@app.route('/api/materials/all', methods=['GET'])
def get_all_materials():
    """Get all materials from database"""
    if not material_service:
        return jsonify({"error": "Database connection not available"}), 500
    
    try:
        print("🔄 Fetching all materials from database...")
        materials = material_service.get_all_materials()  # Get ALL materials without limit
        print(f"✅ Retrieved {len(materials)} materials from database")
        return jsonify(materials)
    except Exception as e:
        print(f"❌ Error fetching materials: {e}")
        return jsonify({"error": f"Failed to fetch materials: {str(e)}"}), 500

@app.route('/api/materials/template', methods=['GET'])
def download_excel_template():
    """Download Excel template for material import"""
    try:
        # Create template data with all columns
        template_data = {
            'material_name': ['混凝土 (範例)', '鋼筋 (範例)', ''],
            'carbon_footprint': [320.5, 1850.0, ''],
            'declaration_unit': ['kg/m³', 'kg/kg', ''],
            'data_source': ['環保署資料 (選填)', 'ISO標準 (選填)', ''],
            'announcement_year': [2023, 2022, ''],
            'life_cycle_scope': ['A1-A3', 'A1-A5', ''],
            'verified': ['是', '否', ''],
            'remarks': ['備註說明 (選填)', '另一個範例', '']
        }
        
        df = pd.DataFrame(template_data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data to Excel
            df.to_excel(writer, index=False, sheet_name='材料匯入範本')
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['材料匯入範本']
            
            # Define styles
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            # Header style - required columns
            required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            required_font = Font(bold=True, color="CC0000")
            
            # Header style - optional columns  
            optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            optional_font = Font(bold=True, color="0066CC")
            
            # Apply styles to headers
            required_columns = ['material_name', 'carbon_footprint', 'declaration_unit']
            
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                if column in required_columns:
                    cell.fill = required_fill
                    cell.font = required_font
                else:
                    cell.fill = optional_fill  
                    cell.font = optional_font
                    
                # Auto-adjust column width
                worksheet.column_dimensions[cell.column_letter].width = 20
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='材料匯入範本.xlsx'
        )
        
    except Exception as e:
        print(f"Error creating template: {e}")
        return jsonify({"error": f"Failed to create template: {str(e)}"}), 500

@app.route('/api/materials/preview-excel', methods=['POST'])
def preview_excel_materials():
    """Preview Excel file contents before import"""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Please provide a valid Excel file"}), 400
    
    try:
        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['material_name', 'carbon_footprint', 'declaration_unit']
        optional_columns = ['data_source', 'announcement_year', 'life_cycle_scope', 'verified', 'remarks']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                "error": f"Missing required columns: {', '.join(missing_columns)}. Required: {', '.join(required_columns)}"
            }), 400
        
        # Process and validate data
        preview_data = []
        validation_errors = []
        
        for index, row in df.iterrows():
            row_data = {}
            row_errors = []
            
            try:
                # Required fields
                row_data['material_name'] = str(row['material_name']).strip()
                if not row_data['material_name'] or row_data['material_name'].lower() == 'nan':
                    row_errors.append("Material name cannot be empty")
                
                try:
                    row_data['carbon_footprint'] = float(row['carbon_footprint'])
                    if row_data['carbon_footprint'] < 0:
                        row_errors.append("Carbon footprint must be non-negative")
                except (ValueError, TypeError):
                    row_errors.append("Invalid carbon footprint value")
                    row_data['carbon_footprint'] = None
                
                row_data['declaration_unit'] = str(row['declaration_unit']).strip()
                if not row_data['declaration_unit'] or row_data['declaration_unit'].lower() == 'nan':
                    row_errors.append("Declaration unit cannot be empty")
                
                # Optional fields
                for field in optional_columns:
                    if field in df.columns and pd.notna(row[field]):
                        if field == 'announcement_year':
                            try:
                                row_data[field] = int(float(row[field]))
                                if row_data[field] < 1900 or row_data[field] > 2100:
                                    row_errors.append("Invalid announcement year")
                            except (ValueError, TypeError):
                                row_errors.append("Invalid announcement year format")
                                row_data[field] = None
                        else:
                            row_data[field] = str(row[field]).strip()
                    else:
                        row_data[field] = ''
                
                # Add row index and validation status
                row_data['row_index'] = index + 2  # Excel rows start at 2 (accounting for header)
                row_data['is_valid'] = len(row_errors) == 0
                row_data['errors'] = row_errors
                
                preview_data.append(row_data)
                
                if row_errors:
                    validation_errors.extend([f"Row {index + 2}: {error}" for error in row_errors])
                    
            except Exception as e:
                validation_errors.append(f"Row {index + 2}: Unexpected error - {str(e)}")
                preview_data.append({
                    'row_index': index + 2,
                    'is_valid': False,
                    'errors': [f"Unexpected error: {str(e)}"],
                    **{col: '' for col in required_columns + optional_columns}
                })
        
        # Calculate statistics
        valid_count = sum(1 for item in preview_data if item['is_valid'])
        invalid_count = len(preview_data) - valid_count
        
        return jsonify({
            "preview_data": preview_data,
            "total_rows": len(preview_data),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "validation_errors": validation_errors[:20],  # Limit to first 20 errors
            "columns": {
                "required": required_columns,
                "optional": optional_columns
            }
        })
        
    except Exception as e:
        print(f"Error previewing Excel: {e}")
        return jsonify({"error": f"Failed to preview Excel file: {str(e)}"}), 500

@app.route('/api/materials/import-excel', methods=['POST'])
def import_materials_from_excel():
    """Import materials from previewed Excel data"""
    if not material_service:
        return jsonify({"error": "Database connection not available"}), 500
    
    try:
        data = request.get_json()
        if not data or 'materials' not in data:
            return jsonify({"error": "No materials data provided"}), 400
        
        materials_data = data['materials']
        if not isinstance(materials_data, list):
            return jsonify({"error": "Materials data must be an array"}), 400
        
        # Filter only valid materials
        valid_materials = [material for material in materials_data if material.get('is_valid', False)]
        
        if not valid_materials:
            return jsonify({"error": "No valid materials to import"}), 400
        
        # Import materials
        imported_count = 0
        error_count = 0
        errors = []
        
        for material in valid_materials:
            try:
                # Prepare material data for database
                material_data = {
                    'material_name': material['material_name'],
                    'carbon_footprint': material['carbon_footprint'],
                    'declaration_unit': material['declaration_unit'],
                }
                
                # Add optional fields if they exist and are not empty
                optional_fields = ['data_source', 'announcement_year', 'life_cycle_scope', 'verified', 'remarks']
                for field in optional_fields:
                    if field in material and material[field] and str(material[field]).strip():
                        if field == 'announcement_year':
                            material_data[field] = int(material[field])
                        else:
                            material_data[field] = str(material[field]).strip()
                
                # Create material in database
                material_service.create_material(material_data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                row_index = material.get('row_index', 'unknown')
                errors.append(f"Row {row_index}: {str(e)}")
        
        return jsonify({
            "message": f"Import completed. {imported_count} materials imported, {error_count} errors.",
            "imported_count": imported_count,
            "error_count": error_count,
            "errors": errors[:10]  # Limit to first 10 errors
        })
        
    except Exception as e:
        print(f"Error importing materials: {e}")
        return jsonify({"error": f"Failed to import materials: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=8001, host='0.0.0.0')