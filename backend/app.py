# work_station/backend/app.py

# --- 基礎與 Web 套件 ---
import os
import sys
import time
import datetime
import io
import zipfile
import re
import shutil
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from flask_restx import Api, Resource, fields, reqparse
import werkzeug.utils
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# --- 功能性套件 ---
try:
    import googlemaps
except ImportError as e:
    print(f"Warning: {e}. Some features may not work.")
    googlemaps = None

try:
    from supabase_client import supabase
except ImportError as e:
    print(f"Warning: {e}. Some features may not work.")
    supabase = None

try:
    from gmap_robot import GoogleMapsRobot
except ImportError as e:
    print(f"Warning: {e}. Some features may not work.")
    GoogleMapsRobot = None

# --- OCR 相關套件 (延遲導入以加快啟動) ---
# Import these only when OCR功能 is actually needed
# This prevents blocking the app startup with large model downloads
try:
    import cv2
    import numpy as np
    from pdf2image import convert_from_path
    # 從原本的 OCR 工具導入設定變數
    # 請確保 param.py 與 app.py 在同一個資料夾中
    from param import *
    OCR_IMPORTS_AVAILABLE = True
except ImportError as e:
    print(f"Warning: OCR dependencies not fully available: {e}")
    OCR_IMPORTS_AVAILABLE = False

# Lazy import OCR engines - these will be imported only when init_ocr_engines() is called
CnOcr = None
easyocr = None
PaddleOCR = None

# --- 解決 Flask 在 Windows 中 print() 可能產生的亂碼問題 ---
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ======================================================================
# --- 應用程式初始化與設定 (只需一次) ---
# ======================================================================
app = Flask(__name__)
CORS(app, 
     origins=['http://localhost:5173', 'http://localhost:5174', 'http://127.0.0.1:5173', 'http://127.0.0.1:5174'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization'])

# Swagger/OpenAPI configuration
api = Api(
    app,
    version='1.0',
    title='MFish Station Backend API',
    description='Backend API for MFish Station - OCR, Google Maps, and Material Management',
    doc='/docs/',
    prefix='/api'
)

# --- 路徑設定 ---
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SCREENSHOTS_FOLDER'] = os.path.join(basedir, 'screenshots')
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'uploads')
app.config['REPORTS_FOLDER'] = os.path.join(basedir, 'reports')
app.config['TEMP_IMG_FOLDER'] = os.path.join(basedir, 'temp_imgs')
app.config['CROPPED_RECEIPTS_FOLDER'] = os.path.join(basedir, 'cropped_receipts')

# --- 全域物件 ---
GOOGLE_MAPS_API_KEY = os.getenv("MAPS_API_KEY")
gmaps = googlemaps.Client(key=GOOGLE_MAPS_API_KEY) if GOOGLE_MAPS_API_KEY and googlemaps else None
SESSION_RESULTS_CACHE = {}

# OCR 引擎 (延遲初始化)
ocr_engines = { "cnocr": None, "easyocr": None, "paddleocr": None }

# ======================================================================
# --- OCR 核心邏輯 (100% 移植自 gas_helper.py) ---
# ======================================================================

def init_ocr_engines():
    """初始化所有 OCR 引擎。"""
    global CnOcr, easyocr, PaddleOCR

    if ocr_engines["cnocr"] is None:
        print("首次使用，正在初始化 OCR 引擎 (可能需要幾分鐘)...")

        # Lazy import OCR libraries here (not at module level)
        if CnOcr is None:
            try:
                from cnocr import CnOcr as CnOcrClass
                CnOcr = CnOcrClass
                print("CnOCR library imported successfully")
            except ImportError as e:
                print(f"Failed to import CnOCR: {e}")

        if easyocr is None:
            try:
                import easyocr as easyocr_module
                easyocr = easyocr_module
                print("EasyOCR library imported successfully")
            except ImportError as e:
                print(f"Failed to import EasyOCR: {e}")

        if PaddleOCR is None:
            try:
                from paddleocr import PaddleOCR as PaddleOCRClass
                PaddleOCR = PaddleOCRClass
                print("PaddleOCR library imported successfully")
            except ImportError as e:
                print(f"Failed to import PaddleOCR: {e}")

        # 初始化 CnOcr - 使用 try/except 處理可能的導入問題
        if CnOcr:
            try:
                ocr_engines["cnocr"] = CnOcr()
                print("CnOCR 初始化成功！")
            except Exception as e:
                print(f"CnOCR 初始化失敗: {e}")
                ocr_engines["cnocr"] = None
        else:
            print("CnOCR not available, skipping")

        # 初始化 EasyOCR
        if easyocr:
            try:
                ocr_engines["easyocr"] = easyocr.Reader(['ch_tra', 'en'])
                print("EasyOCR 初始化成功！")
            except Exception as e:
                print(f"EasyOCR 初始化失敗: {e}")
                ocr_engines["easyocr"] = None
        else:
            print("EasyOCR not available, skipping")

        # 初始化 PaddleOCR - 處理版本兼容性問題
        if PaddleOCR:
            try:
                # 先嘗試不使用已棄用的參數
                ocr_engines["paddleocr"] = PaddleOCR(use_textline_orientation=False, lang='ch')
                print("PaddleOCR 初始化成功！")
            except Exception as e:
                print(f"PaddleOCR 初始化失敗 (新版): {e}")
                try:
                    # 回退到舊版參數
                    ocr_engines["paddleocr"] = PaddleOCR(use_angle_cls=False, lang='ch')
                    print("PaddleOCR 初始化成功 (舊版)！")
                except Exception as e2:
                    print(f"PaddleOCR 完全初始化失敗: {e2}")
                    ocr_engines["paddleocr"] = None
        else:
            print("PaddleOCR not available, skipping")

        print("OCR 引擎初始化完成！")

def detect_invoices_from_pdf(pdf_path: str) -> list:
    """從 PDF 中分割出所有發票圖片。"""
    os.makedirs(app.config['TEMP_IMG_FOLDER'], exist_ok=True)
    os.makedirs(app.config['CROPPED_RECEIPTS_FOLDER'], exist_ok=True)
    invoice_images = []
    images = convert_from_path(pdf_path, dpi=300)
    for i, img in enumerate(images):
        page_filename = f'page_{i + 1}.png'
        img_path = os.path.join(app.config['TEMP_IMG_FOLDER'], page_filename)
        img.save(img_path, 'PNG')
        image = cv2.imread(img_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        bin_img = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 25, 15)
        height, width = gray.shape[:2]
        scale = max(width, height) / 1000
        ksize = int(30 * scale)
        ksize = max(20, min(ksize, 80))
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (ksize, ksize))
        dilated = cv2.dilate(bin_img, kernel, iterations=1)
        contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        boxes = [cv2.boundingRect(c) for c in contours if cv2.contourArea(c) > 5000]
        boxes = sorted(boxes, key=lambda b: (b[1], b[0]))
        for j, (x, y, w, h) in enumerate(boxes):
            crop_img = image[y:y + h, x:x + w]
            crop_path = os.path.join(app.config['CROPPED_RECEIPTS_FOLDER'], f'{page_filename}_block{j}.png')
            cv2.imwrite(crop_path, crop_img)
            invoice_images.append(crop_path)
    return invoice_images

def detect_fuel_type(text_combined: str) -> str:
    """從文字中偵測燃油種類。"""
    matches = []
    for wrong, correct in fuel_fuzzy_mapping.items():
        text_combined = text_combined.replace(wrong, correct)
    for fuel in fuel_keywords:
        if fuel in text_combined:
            mapped = fuel_mapping.get(fuel, fuel)
            matches.append((fuel, mapped))
    if matches:
        return sorted(matches, key=lambda x: -len(x[0]))[0][1]
    return None

def extract_invoice_info(img_path: str) -> dict:
    """從單張發票圖片中擷取資訊 (完整版)。"""
    cnocr_lines = [''.join(block['text']) for block in ocr_engines["cnocr"].ocr(img_path)]
    zh_lines = ocr_engines["easyocr"].readtext(img_path, detail=0)
    all_lines = cnocr_lines + zh_lines
    invoice_number, date, quantity, fuel_type, address = None, None, None, None, None

    for line in cnocr_lines:
        if not invoice_number and (match := invoice_number_pattern.search(line)): invoice_number = match.group()
        if not date and (match := date_pattern.search(line)): date = match.group()

    for i, line in enumerate(all_lines):
        if not quantity and any(k in line for k in fuel_keywords):
            if (match := quantity_pattern.search(line)): quantity = match.group(1); break
            if (match := quantity_fallback_pattern.search(line)): quantity = match.group(); break
            if i + 1 < len(all_lines):
                next_line = all_lines[i + 1]
                if (match := quantity_pattern.search(next_line)): quantity = match.group(1); break
                if (match := quantity_fallback_pattern.search(next_line)): quantity = match.group(); break

    if not quantity:
        for line in all_lines:
            if (match := quantity_pattern.search(line)): quantity = match.group(1); break

    all_text_combined = ' '.join(all_lines)
    fuel_type = detect_fuel_type(all_text_combined)

    for line in zh_lines:
        if not address and (match := address_pattern.search(line)): address = line; break
    if not address:
        for line in zh_lines:
            if any(keyword in line for keyword in district_keywords) and '號' in line and re.search(r'\d+', line): address = line; break

    # --- 關鍵：完整的 PaddleOCR 備用方案 ---
    if not all([invoice_number, date, quantity, fuel_type, address]) and ocr_engines["paddleocr"] is not None:
        paddle_result = ocr_engines["paddleocr"].ocr(img_path) # 修正：移除 cls=False
        if paddle_result and paddle_result[0]:
            paddle_lines = [line[1][0] for line in paddle_result[0]]
            if not invoice_number:
                for line in paddle_lines:
                    if (match := invoice_number_pattern.search(line)): invoice_number = match.group(); break
            if not date:
                for line in paddle_lines:
                    if (match := date_pattern.search(line)): date = match.group(); break
            if not quantity:
                for line in paddle_lines:
                    if (match := simple_quantity_pattern.search(line)): quantity = match.group(1); break
            if not fuel_type:
                text_joined = ' '.join(paddle_lines)
                fuel_type = detect_fuel_type(text_joined)

    # --- 關鍵：完整的資料清理 ---
    if address:
        replacements = {'半禹锈娜': '萬巒鄉', '号': '號', '锈': '', '娜': '', '潮洲': '潮州', '川': '州', '鎖': '鎮'}
        for wrong, right in replacements.items(): address = address.replace(wrong, right)
    if invoice_number and not invoice_number_pattern.fullmatch(invoice_number): invoice_number = None
    if date and not date_pattern.fullmatch(date): date = None
    if quantity:
        try: float(quantity)
        except ValueError: quantity = None
    if fuel_type and fuel_type not in fuel_mapping.values(): fuel_type = None
    if address and (len(address) < 6 or '號' not in address or not any(city in address for city in district_keywords)): address = None

    print(f"  > OCR 結果: {invoice_number}, {date}, {fuel_type}, {quantity}, {address}")
    return {
        '頁數': os.path.basename(img_path), '發票號碼': invoice_number, '日期': date,
        '種類': fuel_type, '數量': quantity, '地址': address, '備註': ''
    }

def process_single_invoice_thread_safe(img_path: str, thread_id: int) -> dict:
    """線程安全的單張發票處理函數。"""
    try:
        print(f"  🧵 Thread {thread_id}: Processing {os.path.basename(img_path)}")
        result = extract_invoice_info(img_path)
        print(f"  ✅ Thread {thread_id}: Completed {os.path.basename(img_path)} - Invoice: {result.get('發票號碼', 'None')}")
        return result
    except Exception as e:
        print(f"  ❌ Thread {thread_id}: Error processing {os.path.basename(img_path)}: {e}")
        return {
            '頁數': os.path.basename(img_path), '發票號碼': None, '日期': None,
            '種類': None, '數量': None, '地址': None, '備註': f'處理錯誤: {str(e)}'
        }

def process_invoice_pdf(pdf_path: str) -> tuple[str, list]:
    """整合的多線程 OCR 處理流程，回傳報告路徑和結果資料。"""
    init_ocr_engines()
    print(f"正在處理 PDF: {pdf_path}")
    invoice_images = detect_invoices_from_pdf(pdf_path)
    num_invoices = len(invoice_images)
    print(f"分割出 {num_invoices} 張發票，開始多線程 OCR 辨識...")

    # 動態決定線程數量：最多4個線程，但不超過發票數量
    max_workers = min(4, num_invoices, os.cpu_count() or 4)
    print(f"🚀 使用 {max_workers} 個線程進行並行處理...")

    results = [None] * num_invoices  # 預分配結果列表以保持順序
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任務
        future_to_index = {
            executor.submit(process_single_invoice_thread_safe, img_path, i + 1): i
            for i, img_path in enumerate(invoice_images)
        }

        # 收集結果並保持原始順序
        completed_count = 0
        for future in as_completed(future_to_index):
            index = future_to_index[future]
            results[index] = future.result()
            completed_count += 1
            progress = (completed_count / num_invoices) * 100
            print(f"  📊 進度: {completed_count}/{num_invoices} ({progress:.1f}%)")

    processing_time = time.time() - start_time
    print(f"🎉 多線程處理完成！耗時: {processing_time:.2f}秒 (平均: {processing_time/num_invoices:.2f}秒/張)")

    df = pd.DataFrame(results)
    report_filename = f'ocr_report_{int(time.time())}.xlsx'
    report_path = os.path.join(app.config['REPORTS_FOLDER'], report_filename)
    df.to_excel(report_path, index=False)
    print(f"報告已產生: {report_path}")
    if os.path.exists(app.config['TEMP_IMG_FOLDER']): shutil.rmtree(app.config['TEMP_IMG_FOLDER'])
    if os.path.exists(app.config['CROPPED_RECEIPTS_FOLDER']): shutil.rmtree(app.config['CROPPED_RECEIPTS_FOLDER'])
    return report_path, results

# ======================================================================
# --- 輔助函式 ---
# ======================================================================
def get_origin_city(origin):
    """從完整的出發地地址中，提取出縣市名稱。"""
    city_keywords = ["台北市", "新北市", "桃園市", "台中市", "台南市", "高雄市", "基隆市", "新竹市", "嘉義市", "宜蘭縣", "新竹縣", "苗栗縣", "彰化縣", "南投縣", "雲林縣", "嘉義縣", "屏東縣", "花蓮縣", "台東縣", "澎湖縣", "金門縣", "連江縣"]
    for city in city_keywords:
        if city in origin:
            return city
    return ""

# ======================================================================
# --- API Models for Swagger Documentation ---
# ======================================================================

# API Models
hello_model = api.model('HelloResponse', {
    'message': fields.String(required=True, description='Response message')
})

material_model = api.model('Material', {
    'material_id': fields.String(required=True, description='Material ID'),
    'material_name': fields.String(required=True, description='Material name'),
    'carbon_footprint': fields.Float(required=True, description='Carbon footprint value'),
    'declaration_unit': fields.String(required=True, description='Declaration unit'),
    'data_source': fields.String(description='Data source'),
    'life_cycle_scope': fields.String(description='Life cycle scope'),
    'announcement_year': fields.Integer(description='Announcement year'),
    'verified': fields.String(description='Verified by'),
    'remarks': fields.String(description='Remarks')
})

material_match_model = api.model('MaterialMatch', {
    'name': fields.String(required=True, description='Material name'),
    'id': fields.String(required=True, description='Material ID'),
    'carbon_footprint': fields.Float(required=True, description='Carbon footprint value'),
    'declaration_unit': fields.String(required=True, description='Declaration unit'),
    'score': fields.Float(required=True, description='Match score')
})

material_create_model = api.model('MaterialCreate', {
    'material_name': fields.String(required=True, description='Material name'),
    'carbon_footprint': fields.Float(required=True, description='Carbon footprint value'),
    'declaration_unit': fields.String(required=True, description='Declaration unit'),
    'data_source': fields.String(description='Data source'),
    'life_cycle_scope': fields.String(description='Life cycle scope'),
    'announcement_year': fields.Integer(description='Announcement year'),
    'verified': fields.String(description='Verified by'),
    'remarks': fields.String(description='Remarks')
})

success_response_model = api.model('SuccessResponse', {
    'success': fields.Boolean(required=True, description='Success status'),
    'data': fields.Raw(description='Response data'),
    'message': fields.String(description='Response message')
})

material_batch_result_model = api.model('MaterialBatchResult', {
    'query': fields.String(required=True, description='Original query'),
    'matches': fields.List(fields.Nested(material_match_model)),
    'default': fields.Integer(description='Default selection index')
})

gmap_request_model = api.model('GMapRequest', {
    'origin': fields.String(required=True, description='Starting location'),
    'destinations': fields.String(required=True, description='Destinations (newline separated)')
})

gmap_result_model = api.model('GMapResult', {
    'origin': fields.String(required=True, description='Starting location'),
    'destination': fields.String(required=True, description='Destination'),
    'distance': fields.String(required=True, description='Distance information'),
    'image_filename': fields.String(description='Screenshot filename'),
    'screenshot_url': fields.String(description='Screenshot URL')
})

gmap_response_model = api.model('GMapResponse', {
    'results': fields.List(fields.Nested(gmap_result_model)),
    'session_id': fields.String(required=True, description='Session ID for downloads')
})

ocr_response_model = api.model('OCRResponse', {
    'message': fields.String(required=True, description='Processing status message'),
    'download_url': fields.String(required=True, description='Download URL for Excel report'),
    'data': fields.List(fields.Raw, description='OCR extracted data')
})

error_model = api.model('Error', {
    'error': fields.String(required=True, description='Error message')
})

# API Namespaces
ns_general = api.namespace('general', description='General operations')
ns_materials = api.namespace('materials', description='Material matching operations')
ns_gmap = api.namespace('gmap', description='Google Maps operations')
ns_ocr = api.namespace('ocr', description='OCR processing operations')

# ======================================================================
# --- API 端點 (Endpoints) ---
# ======================================================================

@app.route('/')
def index():
    return jsonify({
        "message": "MFish Station Backend API",
        "status": "running",
        "swagger_docs": "/docs/",
        "endpoints": {
            "hello": "/api/general/hello",
            "materials": "/api/materials/match-batch", 
            "gmap": "/api/gmap/process",
            "ocr": "/api/ocr/process-pdf"
        }
    })

@ns_general.route('/hello')
class HelloWorld(Resource):
    @ns_general.doc('hello_world')
    @ns_general.marshal_with(hello_model)
    def get(self):
        """Test API connection and database status"""
        if supabase: 
            return {"message": "哈囉！我來自成功連線到 Supabase 的 Python 後端！"}
        else: 
            api.abort(500, "資料庫連線失敗")

@ns_general.route('/health')
class HealthCheck(Resource):
    @ns_general.doc('health_check')
    def get(self):
        """Check service health and status"""
        try:
            # Test database connection
            db_status = 'connected'
            if supabase:
                try:
                    # Simple test query to verify connection
                    test_response = supabase.table('materials').select('material_id').limit(1).execute()
                    db_status = 'connected'
                except:
                    db_status = 'disconnected'
            else:
                db_status = 'disconnected'
            
            # Check other services
            gmaps_status = 'available' if GOOGLE_MAPS_API_KEY and gmaps else 'unavailable'
            ocr_status = 'available'  # OCR is commented out but available
            
            return {
                "success": True,
                "data": {
                    "status": "healthy",
                    "timestamp": datetime.now().isoformat(),
                    "services": {
                        "database": db_status,
                        "google_maps": gmaps_status,
                        "ocr": ocr_status
                    }
                }
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Health check failed: {str(e)}"
            }, 500

@ns_general.route('/info')
class ServiceInfo(Resource):
    @ns_general.doc('service_info')
    def get(self):
        """Get service information"""
        return {
            "success": True,
            "data": {
                "name": "MFish Station Backend API",
                "version": "1.0",
                "description": "Backend API for MFish Station - OCR, Google Maps, and Material Management",
                "endpoints": {
                    "health": "/api/general/health",
                    "materials": "/api/materials/",
                    "gmap": "/api/gmap/process",
                    "ocr": "/api/ocr/process-pdf"
                }
            }
        }

@ns_materials.route('/match-batch')
class MaterialMatchBatch(Resource):
    @ns_materials.doc('match_materials_batch')
    @ns_materials.expect(api.model('MaterialQueries', {
        'queries': fields.List(fields.String, required=True, description='List of material names to search')
    }))
    @ns_materials.marshal_with(success_response_model)
    def post(self):
        """Batch match materials against database"""
        if not supabase: 
            api.abort(500, "資料庫連線失敗")
        
        data = request.get_json()
        queries = data.get('queries', []) if data else []
        if not queries: 
            api.abort(400, "沒有收到任何查詢資料")
        
        all_results = []
        try:
            for original_name in queries:
                # 使用正確的欄位名稱進行搜尋
                response = supabase.table('materials').select('material_id, material_name, carbon_footprint, declaration_unit, data_source').ilike('material_name', f'%{original_name}%').limit(5).execute()
                search_results = response.data if response.data else []
                
                # 轉換格式以符合前端期望
                formatted_matches = []
                for material in search_results:
                    formatted_matches.append({
                        "name": material.get('material_name', ''),
                        "id": material.get('material_id', ''),
                        "carbon_footprint": material.get('carbon_footprint', 0),
                        "declaration_unit": material.get('declaration_unit', ''),
                        "data_source": material.get('data_source', ''),
                        "score": 0.8  # 暫時給一個固定分數
                    })
                
                all_results.append({ 
                    "query": original_name,
                    "matches": formatted_matches,
                    "default": 0 if formatted_matches else None
                })
            
            return {
                "success": True,
                "data": all_results
            }
        except Exception as e:
            print(f"批次比對時發生錯誤: {e}")
            api.abort(500, f"批次比對時發生錯誤: {e}")

# Legacy route for backward compatibility
@app.route('/materials/match-batch', methods=['POST'])
def match_materials_batch():
    if not supabase: return jsonify({"success": False, "error": "資料庫連線失敗"}), 500
    data = request.get_json()
    queries = data.get('queries', []) if data else []
    if not queries: return jsonify({"success": False, "error": "沒有收到任何查詢資料"}), 400
    all_results = []
    try:
        for original_name in queries:
            response = supabase.table('materials').select('material_id, material_name, carbon_footprint, declaration_unit, data_source').ilike('material_name', f'%{original_name}%').limit(5).execute()
            search_results = response.data if response.data else []
            formatted_matches = []
            for material in search_results:
                formatted_matches.append({
                    "name": material.get('material_name', ''),
                    "id": material.get('material_id', ''),
                    "carbon_footprint": material.get('carbon_footprint', 0),
                    "declaration_unit": material.get('declaration_unit', ''),
                    "data_source": material.get('data_source', ''),
                    "score": 0.8
                })
            all_results.append({ 
                "query": original_name,
                "matches": formatted_matches,
                "default": 0 if formatted_matches else None
            })
        return jsonify({"success": True, "data": all_results})
    except Exception as e:
        print(f"批次比對時發生錯誤: {e}")
        return jsonify({"success": False, "error": f"批次比對時發生錯誤: {e}"}), 500

@ns_materials.route('/all')
class MaterialsAll(Resource):
    @ns_materials.doc('get_all_materials')
    @ns_materials.marshal_with(success_response_model)
    def get(self):
        """Get all materials from database using the material service"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        try:
            # Import and use the material service
            from services.material_service import MaterialService
            material_service = MaterialService(supabase)
            
            print("🔄 Using MaterialService to fetch all materials...")
            all_materials = material_service.get_all_materials()
            
            return {
                "success": True,
                "data": all_materials,
                "total_count": len(all_materials)
            }
        except Exception as e:
            print(f"獲取所有材料時發生錯誤: {e}")
            api.abort(500, f"獲取材料時發生錯誤: {e}")

@ns_materials.route('/count')
class MaterialsCount(Resource):
    @ns_materials.doc('get_materials_count')
    def get(self):
        """Get total count of materials in database"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        try:
            # Get exact count using Supabase count feature
            response = supabase.table('materials').select('*', count='exact').limit(1).execute()
            
            return {
                "success": True,
                "count": response.count,
                "message": f"資料庫中共有 {response.count} 筆材料記錄"
            }
        except Exception as e:
            print(f"獲取材料數量時發生錯誤: {e}")
            api.abort(500, f"獲取材料數量時發生錯誤: {e}")

@ns_materials.route('/search')
class MaterialsSearch(Resource):
    @ns_materials.doc('search_materials')
    @ns_materials.marshal_with(success_response_model)
    def get(self):
        """Search materials by query"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        parser = reqparse.RequestParser()
        parser.add_argument('q', type=str, required=True, help='Search query')
        parser.add_argument('limit', type=int, default=5, help='Result limit')
        args = parser.parse_args()
        
        query = args['q'].strip()
        limit = args['limit']
        
        if not query:
            api.abort(400, "搜尋查詢不能為空")
        
        try:
            # Search in material_name column with partial matching
            response = supabase.table('materials').select('*').ilike('material_name', f'%{query}%').limit(limit).execute()
            materials = response.data if response.data else []
            
            return {
                "success": True,
                "data": materials
            }
        except Exception as e:
            print(f"搜尋材料時發生錯誤: {e}")
            api.abort(500, f"搜尋材料時發生錯誤: {e}")

@ns_materials.route('')
class Materials(Resource):
    @ns_materials.doc('create_material')
    @ns_materials.expect(material_create_model)
    @ns_materials.marshal_with(success_response_model)
    def post(self):
        """Create a new material"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        data = request.get_json()
        if not data:
            api.abort(400, "請提供材料數據")
        
        # Validate required fields
        required_fields = ['material_name', 'carbon_footprint', 'declaration_unit']
        for field in required_fields:
            if not data.get(field):
                api.abort(400, f"缺少必填字段: {field}")
        
        try:
            # Insert material into database
            material_data = {
                'material_name': data['material_name'],
                'carbon_footprint': float(data['carbon_footprint']),
                'declaration_unit': data['declaration_unit'],
                'data_source': data.get('data_source', ''),
                'life_cycle_scope': data.get('life_cycle_scope', ''),
                'announcement_year': int(data['announcement_year']) if data.get('announcement_year') else None,
                'verified': data.get('verified', ''),
                'remarks': data.get('remarks', '')
            }
            
            response = supabase.table('materials').insert(material_data).execute()
            
            if response.data:
                return {
                    "success": True,
                    "data": response.data[0] if response.data else material_data,
                    "message": "材料創建成功"
                }
            else:
                api.abort(500, "創建材料失敗")
                
        except Exception as e:
            print(f"創建材料時發生錯誤: {e}")
            api.abort(500, f"創建材料時發生錯誤: {e}")

@ns_materials.route('/<string:material_id>')
class MaterialById(Resource):
    @ns_materials.doc('get_material_by_id')
    @ns_materials.marshal_with(success_response_model)
    def get(self, material_id):
        """Get material by ID"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        try:
            response = supabase.table('materials').select('*').eq('material_id', material_id).execute()
            
            if response.data and len(response.data) > 0:
                return {
                    "success": True,
                    "data": response.data[0]
                }
            else:
                api.abort(404, "材料未找到")
                
        except Exception as e:
            print(f"獲取材料時發生錯誤: {e}")
            api.abort(500, f"獲取材料時發生錯誤: {e}")
    
    @ns_materials.doc('update_material')
    @ns_materials.expect(material_create_model)
    @ns_materials.marshal_with(success_response_model)
    def put(self, material_id):
        """Update material by ID"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        data = request.get_json()
        if not data:
            api.abort(400, "請提供更新數據")
        
        try:
            # Prepare update data (only include provided fields)
            update_data = {}
            allowed_fields = ['material_name', 'carbon_footprint', 'declaration_unit', 
                            'data_source', 'life_cycle_scope', 'announcement_year', 'verified', 'remarks']
            
            for field in allowed_fields:
                if field in data:
                    if field == 'carbon_footprint' and data[field] is not None:
                        update_data[field] = float(data[field])
                    elif field == 'announcement_year' and data[field] is not None:
                        update_data[field] = int(data[field])
                    else:
                        update_data[field] = data[field]
            
            if not update_data:
                api.abort(400, "沒有提供有效的更新數據")
            
            response = supabase.table('materials').update(update_data).eq('material_id', material_id).execute()
            
            if response.data and len(response.data) > 0:
                return {
                    "success": True,
                    "data": response.data[0],
                    "message": "材料更新成功"
                }
            else:
                api.abort(404, "材料未找到或更新失敗")
                
        except Exception as e:
            print(f"更新材料時發生錯誤: {e}")
            api.abort(500, f"更新材料時發生錯誤: {e}")
    
    @ns_materials.doc('delete_material')
    @ns_materials.marshal_with(success_response_model)
    def delete(self, material_id):
        """Delete material by ID"""
        if not supabase:
            api.abort(500, "資料庫連線失敗")
        
        try:
            # Check if material exists first
            check_response = supabase.table('materials').select('material_id').eq('material_id', material_id).execute()
            
            if not check_response.data or len(check_response.data) == 0:
                api.abort(404, "材料未找到")
            
            # Delete the material
            response = supabase.table('materials').delete().eq('material_id', material_id).execute()
            
            return {
                "success": True,
                "data": None,
                "message": "材料刪除成功"
            }
                
        except Exception as e:
            print(f"刪除材料時發生錯誤: {e}")
            api.abort(500, f"刪除材料時發生錯誤: {e}")

# ============================================================================
# Excel Import/Export Endpoints (Plain Flask Routes)
# ============================================================================

@app.route('/api/materials/template', methods=['GET'])
def download_excel_template():
    """Download Excel template for material import"""
    try:
        # Create template data with all columns
        template_data = {
            'material_name': ['混凝土 (範例)', '鋼筋 (範例)', ''],
            'carbon_footprint': [320.5, 1850.0, ''],
            'declaration_unit': ['kg/m³', 'kg/kg', ''],
            'data_source': ['環保署資料 (選填)', 'ISO標準 (選填)', ''],
            'announcement_year': [2023, 2022, ''],
            'life_cycle_scope': ['A1-A3', 'A1-A5', ''],
            'verified': ['是', '否', ''],
            'remarks': ['備註說明 (選填)', '另一個範例', '']
        }

        df = pd.DataFrame(template_data)
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data to Excel
            df.to_excel(writer, index=False, sheet_name='材料匯入範本')

            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['材料匯入範本']

            # Define styles
            from openpyxl.styles import Font, PatternFill, Border, Side

            # Header style - required columns
            required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            required_font = Font(bold=True, color="CC0000")

            # Header style - optional columns
            optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            optional_font = Font(bold=True, color="0066CC")

            # Apply styles to headers
            required_columns = ['material_name', 'carbon_footprint', 'declaration_unit']

            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                if column in required_columns:
                    cell.fill = required_fill
                    cell.font = required_font
                else:
                    cell.fill = optional_fill
                    cell.font = optional_font

                # Auto-adjust column width
                worksheet.column_dimensions[cell.column_letter].width = 20

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='材料匯入範本.xlsx'
        )

    except Exception as e:
        print(f"Error creating template: {e}")
        return jsonify({"error": f"Failed to create template: {str(e)}"}), 500

@app.route('/api/materials/preview-excel', methods=['POST'])
def preview_excel_materials():
    """Preview Excel file contents before import"""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Please provide a valid Excel file"}), 400

    try:
        # Windows compatibility fix: Read file into memory
        print(f"Reading Excel file: {file.filename}")
        file_content = file.read()
        file.seek(0)  # Reset pointer

        # Create BytesIO object
        excel_buffer = io.BytesIO(file_content)

        # Determine engine based on file extension
        filename = file.filename.lower()
        engine = 'openpyxl' if filename.endswith('.xlsx') else 'xlrd'

        print(f"Using pandas engine: {engine}")

        # Read Excel file with explicit engine
        df = pd.read_excel(excel_buffer, engine=engine)

        print(f"Successfully read {len(df)} rows from Excel")

        # Validate required columns
        required_columns = ['material_name', 'carbon_footprint', 'declaration_unit']
        optional_columns = ['data_source', 'announcement_year', 'life_cycle_scope', 'verified', 'remarks']

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                "error": f"Missing required columns: {', '.join(missing_columns)}. Required: {', '.join(required_columns)}"
            }), 400

        # Process and validate data
        preview_data = []
        validation_errors = []

        for index, row in df.iterrows():
            row_data = {}
            row_errors = []

            try:
                # Required fields
                row_data['material_name'] = str(row['material_name']).strip()
                if not row_data['material_name'] or row_data['material_name'].lower() == 'nan':
                    row_errors.append("Material name cannot be empty")

                try:
                    row_data['carbon_footprint'] = float(row['carbon_footprint'])
                    if row_data['carbon_footprint'] < 0:
                        row_errors.append("Carbon footprint must be non-negative")
                except (ValueError, TypeError):
                    row_errors.append("Invalid carbon footprint value")
                    row_data['carbon_footprint'] = None

                row_data['declaration_unit'] = str(row['declaration_unit']).strip()
                if not row_data['declaration_unit'] or row_data['declaration_unit'].lower() == 'nan':
                    row_errors.append("Declaration unit cannot be empty")

                # Optional fields
                for field in optional_columns:
                    if field in df.columns and pd.notna(row[field]):
                        if field == 'announcement_year':
                            try:
                                row_data[field] = int(float(row[field]))
                                if row_data[field] < 1900 or row_data[field] > 2100:
                                    row_errors.append("Invalid announcement year")
                            except (ValueError, TypeError):
                                row_errors.append("Invalid announcement year format")
                                row_data[field] = None
                        else:
                            row_data[field] = str(row[field]).strip()
                    else:
                        row_data[field] = ''

                # Add row index and validation status
                row_data['row_index'] = index + 2  # Excel rows start at 2 (accounting for header)
                row_data['is_valid'] = len(row_errors) == 0
                row_data['errors'] = row_errors

                preview_data.append(row_data)

                if row_errors:
                    validation_errors.extend([f"Row {index + 2}: {error}" for error in row_errors])

            except Exception as e:
                error_msg = f"Row {index + 2}: Unexpected error - {str(e)}"
                print(f"Error processing row {index + 2}: {e}")
                validation_errors.append(error_msg)
                preview_data.append({
                    'row_index': index + 2,
                    'is_valid': False,
                    'errors': [f"Unexpected error: {str(e)}"],
                    **{col: '' for col in required_columns + optional_columns}
                })

        # Calculate statistics
        valid_count = sum(1 for item in preview_data if item['is_valid'])
        invalid_count = len(preview_data) - valid_count

        print(f"Preview complete: {valid_count} valid, {invalid_count} invalid rows")

        return jsonify({
            "preview_data": preview_data,
            "total_rows": len(preview_data),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "validation_errors": validation_errors[:20],  # Limit to first 20 errors
            "columns": {
                "required": required_columns,
                "optional": optional_columns
            }
        })

    except Exception as e:
        error_msg = f"Failed to preview Excel file: {str(e)}"
        print(f"Error previewing Excel: {e}")
        import traceback
        traceback.print_exc()  # Print full traceback for debugging
        return jsonify({"error": error_msg}), 500

@app.route('/api/materials/import-excel', methods=['POST'])
def import_materials_from_excel():
    """Import materials from previewed Excel data"""
    if not supabase:
        return jsonify({"error": "Database connection not available"}), 500

    try:
        # Import and instantiate MaterialService
        from services.material_service import MaterialService
        material_service = MaterialService(supabase)

        data = request.get_json()
        if not data or 'materials' not in data:
            return jsonify({"error": "No materials data provided"}), 400

        materials_data = data['materials']
        if not isinstance(materials_data, list):
            return jsonify({"error": "Materials data must be an array"}), 400

        # Filter only valid materials
        valid_materials = [material for material in materials_data if material.get('is_valid', False)]

        if not valid_materials:
            return jsonify({"error": "No valid materials to import"}), 400

        # Import materials
        imported_count = 0
        error_count = 0
        errors = []

        for material in valid_materials:
            try:
                # Prepare material data for database
                material_data = {
                    'material_name': material['material_name'],
                    'carbon_footprint': material['carbon_footprint'],
                    'declaration_unit': material['declaration_unit'],
                }

                # Add optional fields if they exist and are not empty
                optional_fields = ['data_source', 'announcement_year', 'life_cycle_scope', 'verified', 'remarks']
                for field in optional_fields:
                    if field in material and material[field] and str(material[field]).strip():
                        if field == 'announcement_year':
                            material_data[field] = int(material[field])
                        else:
                            material_data[field] = str(material[field]).strip()

                # Create material in database
                material_service.create_material(material_data)
                imported_count += 1

            except Exception as e:
                error_count += 1
                row_index = material.get('row_index', 'unknown')
                errors.append(f"Row {row_index}: {str(e)}")

        return jsonify({
            "message": f"Import completed. {imported_count} materials imported, {error_count} errors.",
            "imported_count": imported_count,
            "error_count": error_count,
            "errors": errors[:10]  # Limit to first 10 errors
        })

    except Exception as e:
        print(f"Error importing materials: {e}")
        return jsonify({"error": f"Failed to import materials: {str(e)}"}), 500

@ns_gmap.route('/process')
class GMapProcess(Resource):
    @ns_gmap.doc('gmap_process')
    @ns_gmap.expect(gmap_request_model)
    @ns_gmap.marshal_with(gmap_response_model)
    def post(self):
        """Process Google Maps routes and generate screenshots"""
        data = request.get_json()
        origin, destinations_text = data.get('origin'), data.get('destinations', '')
        if not origin or not destinations_text: 
            api.abort(400, "必須提供出發地和目的地。")
        
        destinations = [addr.strip() for addr in destinations_text.split('\n') if addr.strip()]
        session_id = f"session_{int(time.time())}"
        today_str = datetime.now().strftime("%Y-%m-%d")
        image_folder_path = os.path.join(app.config['SCREENSHOTS_FOLDER'], today_str, session_id)
        os.makedirs(image_folder_path, exist_ok=True)
        
        try:
            # 使用Google Maps機器人
            robot = GoogleMapsRobot(headless=True)
            robot_results = robot.process_multiple_routes(origin, destinations, image_folder_path)
            
            # 轉換結果格式以符合前端期望
            results = []
            for robot_result in robot_results:
                if "image_filename" in robot_result:
                    screenshot_url = f"screenshots/{today_str}/{session_id}/{robot_result['image_filename']}"
                else:
                    screenshot_url = ""
                
                results.append({
                    "origin": robot_result["origin"],
                    "destination": robot_result["destination"], 
                    "distance": robot_result["distance"],
                    "image_filename": robot_result.get("image_filename", ""),
                    "image_local_path": robot_result.get("image_local_path", ""),
                    "screenshot_url": screenshot_url
                })
            
            SESSION_RESULTS_CACHE[session_id] = results
            return {"results": results, "session_id": session_id}
            
        except Exception as e:
            print(f"處理 Google Maps 機器人時發生嚴重錯誤: {e}")
            api.abort(500, f"處理時發生嚴重錯誤: {e}")

@app.route('/api/download/excel/<session_id>', methods=['GET'])
def download_excel(session_id):
    session_data = SESSION_RESULTS_CACHE.get(session_id)
    if not session_data: return "Session not found or expired.", 404
    export_data = [{"起始點": item['origin'], "終點": item['destination'], "距離": item['distance'], "圖片名稱": item['image_filename']} for item in session_data]
    df = pd.DataFrame(export_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer: df.to_excel(writer, index=False, sheet_name='路線距離報告')
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'google_maps_report_{session_id}.xlsx')

@app.route('/api/download/zip/<session_id>', methods=['GET'])
def download_zip(session_id):
    session_data = SESSION_RESULTS_CACHE.get(session_id)
    if not session_data: return "Session not found or expired.", 404
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in session_data: zf.write(item['image_local_path'], arcname=item['image_filename'])
    memory_file.seek(0)
    return send_file(memory_file, mimetype='application/zip', as_attachment=True, download_name=f'map_images_{session_id}.zip')

@app.route('/screenshots/<path:path>')
def send_screenshot(path):
    return send_from_directory(app.config['SCREENSHOTS_FOLDER'], path)

@ns_ocr.route('/process-pdf')
class OCRProcessPDF(Resource):
    @ns_ocr.doc('ocr_process_pdf')
    @ns_ocr.expect(api.parser().add_argument('file', location='files', type='file', required=True, help='PDF file to process'))
    @ns_ocr.marshal_with(ocr_response_model)
    def post(self):
        """Process PDF file with OCR to extract invoice information"""
        if 'file' not in request.files: 
            api.abort(400, "沒有找到檔案部分")
        
        file = request.files['file']
        if file.filename == '' or not file.filename.lower().endswith('.pdf'): 
            api.abort(400, "未選擇或非 PDF 檔案")
        
        unique_filename = f"{int(time.time())}_{werkzeug.utils.secure_filename(file.filename)}"
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(pdf_path)
        
        try:
            report_path, ocr_data = process_invoice_pdf(pdf_path)
            report_filename = os.path.basename(report_path)
            return {
                "message": "OCR 處理完成！",
                "download_url": f"api/download/ocr-report/{report_filename}",
                "data": ocr_data,
                "total_invoices": len(ocr_data),
                "report_filename": report_filename
            }
        except Exception as e:
            print(f"OCR 處理時發生錯誤: {e}")
            api.abort(500, f"處理失敗: {str(e)}")
        finally:
            if os.path.exists(pdf_path): 
                os.remove(pdf_path)

@app.route('/api/download/ocr-report/<filename>')
def download_ocr_report(filename):
    return send_from_directory(app.config['REPORTS_FOLDER'], filename, as_attachment=True)

# --- 主程式進入點 ---
if __name__ == '__main__':
    for folder_key in ['SCREENSHOTS_FOLDER', 'UPLOAD_FOLDER', 'REPORTS_FOLDER', 'TEMP_IMG_FOLDER', 'CROPPED_RECEIPTS_FOLDER']:
        folder_path = app.config[folder_key]
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

    # Use PORT environment variable for Railway deployment, fallback to 8001 for local development
    port = int(os.getenv('PORT', 8001))
    # In production (Railway), debug should be False
    debug = os.getenv('FLASK_ENV', 'development') == 'development'

    app.run(debug=debug, port=port, host='0.0.0.0')
